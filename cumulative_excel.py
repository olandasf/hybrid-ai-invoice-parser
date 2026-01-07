"""
Kumuliacinis Excel failų generatorius.
Prideda kiekvieną sąskaitą kaip naują sheet į bendrą Excel failą.
"""

import os
import re
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from generate_excel import get_column_map_for_excel, _finalize_excel_formatting

# Kumuliacinių failų konfigūracija
CUMULATIVE_CONFIG = {
    'directory': 'Excel Suvestinės',
    'filename_pattern': 'Akcizo_Apskaita_{year}.xlsx',
    'max_sheets': 50,  # Maksimalus sheet\'ų skaičius viename faile
    'supplier_patterns': {
        # Dažni tiekėjų šablonai lietuvių importo sektoriuje
        'bsc': ['bsc', 'baltic', 'trading'],
        'vinea': ['vinea', 'vine'],
        'rimi': ['rimi', 'maxima'],
        'selver': ['selver', 'estonia'],
        'distributor': ['distributor', 'distr'],
        'import': ['import', 'imp'],
        'wine': ['wine', 'vynas', 'vino'],
        'beer': ['beer', 'alus', 'brew'],
        'spirits': ['spirits', 'degtine', 'vodka', 'whisky']
    }
}

def _should_use_formula(header: str, original_key: str) -> bool:
    """Nustato ar stulpeliui reikia naudoti formulę."""
    formula_fields = [
        "Suma (€)",                          # quantity * unit_price
        "Suma su nuolaida (€)",             # quantity * unit_price_with_discount
        "Akcizas viso (€)",                 # quantity * excise_per_unit
        "Transportas viso (€)",             # quantity * transport_per_unit
        "Savikaina viso be PVM (€)",        # quantity * cost_wo_vat
        "Savikaina viso su PVM (€)",        # quantity * cost_w_vat
        "Savikaina vnt. be PVM (€)",        # unit_price_with_discount + excise_per_unit + transport_per_unit
        "Savikaina vnt. su PVM (€)",        # cost_wo_vat * 1.21
    ]
    return header in formula_fields

def _get_formula_for_field(header: str, original_key: str, row_idx: int, col_idx: int, headers: list) -> str:
    """Grąžina Excel formulę konkrečiam laukui."""
    
    def get_col_letter_by_header(target_header: str) -> str:
        """Randa stulpelio raidę pagal header pavadinimą."""
        try:
            col_index = headers.index(target_header) + 1
            return get_column_letter(col_index)
        except ValueError:
            return "A"  # Fallback
    
    if header == "Suma (€)":
        # Kiekis * Vnt. kaina
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        price_col = get_col_letter_by_header("Vnt. kaina (€)")
        return f"={qty_col}{row_idx}*{price_col}{row_idx}"
    
    elif header == "Suma su nuolaida (€)":
        # Kiekis * Vnt. kaina su nuolaida
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        price_discount_col = get_col_letter_by_header("Vnt. kaina su nuolaida (€)")
        return f"={qty_col}{row_idx}*{price_discount_col}{row_idx}"
    
    elif header == "Akcizas viso (€)":
        # Kiekis * Akcizas vnt.
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        excise_unit_col = get_col_letter_by_header("Akcizas vnt. (€)")
        return f"={qty_col}{row_idx}*{excise_unit_col}{row_idx}"
    
    elif header == "Transportas viso (€)":
        # Kiekis * Transportas vnt.
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        transport_unit_col = get_col_letter_by_header("Transportas vnt. (€)")
        return f"={qty_col}{row_idx}*{transport_unit_col}{row_idx}"
    
    elif header == "Savikaina vnt. be PVM (€)":
        # Vnt. kaina su nuolaida + Akcizas vnt. + Transportas vnt.
        price_discount_col = get_col_letter_by_header("Vnt. kaina su nuolaida (€)")
        excise_unit_col = get_col_letter_by_header("Akcizas vnt. (€)")
        transport_unit_col = get_col_letter_by_header("Transportas vnt. (€)")
        return f"={price_discount_col}{row_idx}+{excise_unit_col}{row_idx}+{transport_unit_col}{row_idx}"
    
    elif header == "Savikaina vnt. su PVM (€)":
        # Savikaina vnt. be PVM * 1.21
        cost_wo_vat_col = get_col_letter_by_header("Savikaina vnt. be PVM (€)")
        return f"={cost_wo_vat_col}{row_idx}*1.21"
    
    elif header == "Savikaina viso be PVM (€)":
        # Kiekis * Savikaina vnt. be PVM
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        cost_wo_vat_col = get_col_letter_by_header("Savikaina vnt. be PVM (€)")
        return f"={qty_col}{row_idx}*{cost_wo_vat_col}{row_idx}"
    
    elif header == "Savikaina viso su PVM (€)":
        # Kiekis * Savikaina vnt. su PVM
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        cost_w_vat_col = get_col_letter_by_header("Savikaina vnt. su PVM (€)")
        return f"={qty_col}{row_idx}*{cost_w_vat_col}{row_idx}"
    
    return None

class CumulativeExcelManager:
    """Kumuliacinių Excel failų valdymo klasė."""
    
    def __init__(self):
        """Inicializuoja kumuliacinį Excel valdyklę."""
        self.directory = Path(CUMULATIVE_CONFIG['directory'])
        self.directory.mkdir(exist_ok=True)
        self.current_year = datetime.now().year
        
    def get_cumulative_file_path(self, year: Optional[int] = None) -> Path:
        """
        Grąžina kumuliacinio failo kelią.
        
        Args:
            year: Metai (default: dabartiniai)
            
        Returns:
            Failo kelias
        """
        if year is None:
            year = self.current_year
            
        filename = CUMULATIVE_CONFIG['filename_pattern'].format(year=year)
        return self.directory / filename

    def ensure_unique_sheet_name(self, workbook: Workbook, proposed_name: str) -> str:
        """
        Užtikrina unikalų sheet pavadinimą.
        
        Args:
            workbook: Excel workbook objektas
            proposed_name: Siūlomas pavadinimas
            
        Returns:
            Unikalus sheet pavadinimas
        """
        existing_names = workbook.sheetnames
        
        if proposed_name not in existing_names:
            return proposed_name
        
        # Jei vardas egzistuoja, pridedame numerį
        counter = 1
        while True:
            new_name = f"{proposed_name[:27]}_{counter:02d}"
            if new_name not in existing_names:
                return new_name
            counter += 1
            
            # Saugumo ribojimas
            if counter > 99:
                return f"{proposed_name[:25]}_{datetime.now().strftime('%H%M')}"
    
    def create_summary_sheet(self, workbook: Workbook) -> None:
        """
        Sukurta suvestinės sheet su formulėmis.
        
        Args:
            workbook: Excel workbook objektas
        """
        # Patikrinti ar suvestinės sheet jau egzistuoja
        if 'SUVESTINĖ' in workbook.sheetnames:
            return
        
        # Sukurti suvestinės sheet
        summary_sheet = workbook.create_sheet('SUVESTINĖ', 0)  # Pirmas sheet
        
        # Header
        headers = ['Sheet', 'Tiekėjas', 'Data', 'Prekių kiekis', 'Bendra suma', 'Akcizų suma', 'Transportas', 'Savikaina']
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Automatinis stulpelių plotis
        for col in range(1, len(headers) + 1):
            summary_sheet.column_dimensions[get_column_letter(col)].width = 15
    
    def update_summary_sheet(self, workbook: Workbook, sheet_name: str, supplier_code: str, products: List[Dict]) -> None:
        """
        Atnaujina suvestinės sheet su naujais duomenimis, naudojant formules.
        """
        summary_sheet = workbook['SUVESTINĖ']
        next_row = summary_sheet.max_row + 1
        
        # Nustatome eilutę, kurioje yra "IŠ VISO" skaičiavimai
        total_row_in_sheet = len(products) + 2

        # Dinamiškai surandame stulpelių raides pagal jų pavadinimus
        columns_map = get_column_map_for_excel()
        
        def get_col_letter_by_header(target_header: str) -> str:
            for i, (key, header) in enumerate(columns_map, 1):
                if header == target_header:
                    return get_column_letter(i)
            return None

        # Surandame reikiamų stulpelių raides
        qty_col = get_col_letter_by_header('Kiekis (vnt)')
        amount_col = get_col_letter_by_header('Suma su nuolaida (€)')
        excise_col = get_col_letter_by_header('Akcizas viso (€)')
        transport_col = get_col_letter_by_header('Transportas viso (€)')
        cost_col = get_col_letter_by_header('Savikaina viso su PVM (€)')

        # Sukuriame duomenų eilutę su formulėmis
        data = [
            sheet_name,
            supplier_code,
            datetime.now().strftime('%Y-%m-%d'),
            f"='{sheet_name}'!{qty_col}{total_row_in_sheet}" if qty_col else 0,
            f"='{sheet_name}'!{amount_col}{total_row_in_sheet}" if amount_col else 0,
            f"='{sheet_name}'!{excise_col}{total_row_in_sheet}" if excise_col else 0,
            f"='{sheet_name}'!{transport_col}{total_row_in_sheet}" if transport_col else 0,
            f"='{sheet_name}'!{cost_col}{total_row_in_sheet}" if cost_col else 0
        ]
        
        for col, value in enumerate(data, 1):
            cell = summary_sheet.cell(row=next_row, column=col, value=value)
            if isinstance(value, str) and value.startswith('='):
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.00'
        
    def add_invoice_to_cumulative_file(self, products: List[Dict], transport_total: float = 0.0, summary: Dict = {}) -> Tuple[str, str]:
        """
        Prideda sąskaitą į kumuliacinį Excel failą.
        
        Args:
            products: Produktų sąrašas (jau papildytas akcizais ir banderolėmis)
            transport_total: Bendra transporto suma
            summary: Sąskaitos suvestinės informacija (su tiekėjo pavadinimu)
            
        Returns:
            (file_path, sheet_name) - failo kelias ir sheet pavadinimas
        """
        if not products:
            raise ValueError("Produktų sąrašas tuščias")
        
        # Nustatyti tiekėjo pavadinimą iš suvestinės
        supplier_name = summary.get('supplier_name', 'Nežinomas Tiekėjas')
        # Išvalyti neleistinus simbolius iš tiekėjo pavadinimo, kurie netinka sheet pavadinimui
        forbidden_chars = r'[]\/?*:'
        supplier_name_cleaned = "".join(c for c in supplier_name if c not in forbidden_chars)
        logging.info(f"Nustatytas tiekėjo pavadinimas: {supplier_name_cleaned}")
        
        # Gauti failo kelią
        file_path = self.get_cumulative_file_path()
        
        # Užkrauti arba sukurti workbook
        if file_path.exists():
            try:
                workbook = load_workbook(str(file_path))
                logging.info(f"Užkrautas esamas failas: {file_path}")
            except Exception as e:
                logging.error(f"Klaida užkraunant failą: {e}")
                workbook = Workbook()
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
        else:
            workbook = Workbook()
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
            logging.info(f"Sukuriamas naujas failas: {file_path}")
        
        # Sukurti suvestinės sheet jei neegzistuoja
        self.create_summary_sheet(workbook)
        
        # Patikrinti sheet'ų skaičių
        if len(workbook.worksheets) >= CUMULATIVE_CONFIG['max_sheets'] + 1:  # +1 dėl suvestinės
            logging.warning(f"Viršytas maksimalus sheet\'ų skaičius ({CUMULATIVE_CONFIG['max_sheets']})")

        # Generuoti sheet pavadinimą
        date_str = datetime.now().strftime('%m-%d')
        base_name = f"{supplier_name_cleaned[:20]}_{date_str}"
        sheet_name = self.ensure_unique_sheet_name(workbook, base_name)
        logging.info(f"Sukuriamas sheet: {sheet_name}")
        
        # Sukurti naują sheet
        worksheet = workbook.create_sheet(sheet_name)
        
        # Gauti stulpelių žemėlapį (jame jau yra banderolių stulpeliai)
        columns_map = get_column_map_for_excel()
        headers = [header for _, header in columns_map]
        
        # Sukurti header eilutę
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        
        # Pridėti duomenis
        for row_idx, product in enumerate(products, 2):
            for col_idx, (original_key, header) in enumerate(columns_map, 1):
                value = product.get(original_key, "")
                
                # Nustatome formulės ar reikšmės tipą
                if _should_use_formula(header, original_key):
                    formula = _get_formula_for_field(header, original_key, row_idx, col_idx, headers)
                    if formula:
                        worksheet.cell(row=row_idx, column=col_idx, value=formula)
                    else:
                        worksheet.cell(row=row_idx, column=col_idx, value=value)
                else:
                    worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Pridėti sumų eilutę
        sum_row = len(products) + 2
        worksheet.cell(row=sum_row, column=1, value="IŠ VISO").font = Font(bold=True)
        
        # Pridėti SUM formules
        sum_columns = ['Kiekis (vnt)', 'Suma (€)', 'Suma su nuolaida (€)', 'Akcizas viso (€)', 
                      'Transportas viso (€)', 'Savikaina viso be PVM (€)', 'Savikaina viso su PVM (€)']
        
        for col_idx, (original_key, header) in enumerate(columns_map, 1):
            if header in sum_columns:
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}2:{col_letter}{sum_row-1})"
                cell = worksheet.cell(row=sum_row, column=col_idx, value=formula)
                cell.font = Font(bold=True)
                cell.number_format = '#,##0.00'

        # Formatuoti stulpelius
        _finalize_excel_formatting(workbook, worksheet, file_path)
        
        # Atnaujinti suvestinės sheet
        self.update_summary_sheet(workbook, sheet_name, supplier_name_cleaned, products)
        
        # Išsaugoti failą
        try:
            workbook.save(str(file_path))
            logging.info(f"Failas sėkmingai išsaugotas: {file_path}")
            return str(file_path), sheet_name
        except Exception as e:
            logging.error(f"Klaida saugant failą: {e}")
            raise
    
    def _format_worksheet(self, worksheet) -> None:
        """Formatuoja worksheet stulpelius ir eilutes."""
        # Automatinis stulpelių plotis
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Užfriksyti header eilutę
        worksheet.freeze_panes = 'A2'
    
    def get_statistics(self) -> Dict:
        """
        Grąžina kumuliacinių failų statistikas.
        
        Returns:
            Statistikų žodynas
        """
        stats = {
            'files_count': 0,
            'total_sheets': 0,
            'current_year_file': None,
            'files': []
        }
        
        try:
            for file_path in self.directory.glob('*.xlsx'):
                file_stats = {
                    'filename': file_path.name,
                    'path': str(file_path),
                    'size_mb': round(file_path.stat().st_size / (1024 * 1024), 2),
                    'modified': datetime.fromtimestamp(file_path.stat().st_mtime).strftime('%Y-%m-%d %H:%M'),
                    'sheets_count': 0
                }
                
                try:
                    workbook = load_workbook(str(file_path), read_only=True)
                    file_stats['sheets_count'] = len(workbook.worksheets)
                    stats['total_sheets'] += file_stats['sheets_count']
                    workbook.close()
                except:
                    pass
                
                stats['files'].append(file_stats)
                stats['files_count'] += 1
                
                # Patikrinti ar tai dabartinių metų failas
                if str(self.current_year) in file_path.name:
                    stats['current_year_file'] = file_stats
                    
        except Exception as e:
            logging.error(f"Klaida gaunant statistikas: {e}")
        
        return stats

# Globalus kumuliacinio Excel valdiklio objektas
cumulative_excel_manager = CumulativeExcelManager()
