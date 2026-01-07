import pandas as pd
import os
from datetime import datetime, UTC
import logging 
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def get_column_map_for_excel():
    """ Grąžina stulpelių žemėlapį (originalus raktas -> nauja antraštė). """
    return [
        ("name", "Prekės pavadinimas"),
        ("volume", "Tūris (L)"),
        ("abv", "Alk. %"),
        ("quantity", "Kiekis (vnt)"),
        ("unit_price", "Vnt. kaina (€)"),
        ("unit_price_with_discount", "Vnt. kaina su nuolaida (€)"),
        ("discount_percentage", "Nuolaida (%)"),
        ("amount", "Suma (€)"),
        ("amount_with_discount", "Suma su nuolaida (€)"),
        ("excise_category_key", "Akcizo kat. raktas"),
        ("excise_category", "Akcizo kategorija"),
        ("excise_per_unit", "Akcizas vnt. (€)"),
        ("excise_total", "Akcizas viso (€)"),
        ("transport_per_unit", "Transportas vnt. (€)"),
        ("transport_total", "Transportas viso (€)"),
        ("cost_wo_vat", "Savikaina vnt. be PVM (€)"),
        ("cost_w_vat", "Savikaina vnt. su PVM (€)"),
        ("cost_wo_vat_total", "Savikaina viso be PVM (€)"),
        ("cost_w_vat_total", "Savikaina viso su PVM (€)"),
        # Banderolių stulpeliai
        ("banderole_type", "Banderolių tipas"),
        ("banderole_start", "Banderolės nuo"),
        ("banderole_end", "Banderolės iki"),
        ("banderole_count", "Banderolių kiekis"),
        ("tariff_group", "Tarifinės grupės kodas"),
    ]

def generate_excel_file_with_formulas(products_data: list) -> str:
    """Generuoja Excel failą su formulėmis vietoj statinių reikšmių."""
    
    if not products_data:
        logging.warning("GENERATE_EXCEL.PY - gavo tuščią products_data sąrašą.")
        # Sukuriame tuščią workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Akcizo apskaita'
        
        # Pridedame header'ius
        headers = [header for _, header in get_column_map_for_excel()]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        filename = f"akcizo_apskaita_{timestamp}.xlsx"
        os.makedirs('output', exist_ok=True)
        output_path = os.path.join('output', filename)
        wb.save(output_path)
        return output_path

    # Sukuriame workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Akcizo apskaita'
    
    # Gauname stulpelių žemėlapį
    columns_map = get_column_map_for_excel()
    headers = [header for _, header in columns_map]
    
    # Sukuriame header eilutę
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
    
    # Pridedame duomenis
    for row_idx, product in enumerate(products_data, 2):  # Pradedame nuo 2 eilutės
        for col_idx, (original_key, header) in enumerate(columns_map, 1):
            value = product.get(original_key, "")
            
            # Nustatome formulės ar reikšmės tipą
            if _should_use_formula(header, original_key):
                formula = _get_formula_for_field(header, original_key, row_idx, col_idx, headers)
                if formula:
                    ws.cell(row=row_idx, column=col_idx, value=formula)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Pridedame sumų eilutę su formulėmis
    sum_row = len(products_data) + 2
    ws.cell(row=sum_row, column=1, value="Iš viso")
    
    for col_idx, (original_key, header) in enumerate(columns_map, 1):
        if col_idx == 1:  # Pirmasis stulpelis jau užpildytas
            continue
            
        if _should_sum_column(header):
            # Sukuriame SUM formulę
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}2:{col_letter}{sum_row-1})"
            cell = ws.cell(row=sum_row, column=col_idx, value=formula)
            cell.font = Font(bold=True)
        else:
            ws.cell(row=sum_row, column=col_idx, value="")
    
    return _finalize_excel_formatting(wb, ws, output_path=None)

def _should_use_formula(header: str, original_key: str) -> bool:
    """Nustato ar stulpeliui reikia naudoti formulę."""
    formula_fields = [
        "Suma (€)",                          # quantity * unit_price
        "Suma su nuolaida (€)",             # quantity * unit_price_with_discount
        "Akcizas viso (€)",                 # quantity * excise_per_unit
        "Transportas viso (€)",             # quantity * transport_per_unit
        "Savikaina viso be PVM (€)",        # quantity * cost_wo_vat
        "Savikaina viso su PVM (€)",        # quantity * cost_w_vat
        "Savikaina vnt. be PVM (€)",        # unit_price_with_discount + excise_per_unit + transport_per_unit
        "Savikaina vnt. su PVM (€)",        # cost_wo_vat * 1.21
    ]
    return header in formula_fields

def _get_formula_for_field(header: str, original_key: str, row_idx: int, col_idx: int, headers: list) -> str:
    """Grąžina Excel formulę konkrečiam laukui."""
    
    def get_col_letter_by_header(target_header: str) -> str:
        """Randa stulpelio raidę pagal header pavadinimą."""
        try:
            col_index = headers.index(target_header) + 1
            return get_column_letter(col_index)
        except ValueError:
            return "A"  # Fallback
    
    if header == "Suma (€)":
        # Kiekis * Vnt. kaina
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        price_col = get_col_letter_by_header("Vnt. kaina (€)")
        return f"={qty_col}{row_idx}*{price_col}{row_idx}"
    
    elif header == "Suma su nuolaida (€)":
        # Kiekis * Vnt. kaina su nuolaida
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        price_discount_col = get_col_letter_by_header("Vnt. kaina su nuolaida (€)")
        return f"={qty_col}{row_idx}*{price_discount_col}{row_idx}"
    
    elif header == "Akcizas viso (€)":
        # Kiekis * Akcizas vnt.
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        excise_unit_col = get_col_letter_by_header("Akcizas vnt. (€)")
        return f"={qty_col}{row_idx}*{excise_unit_col}{row_idx}"
    
    elif header == "Transportas viso (€)":
        # Kiekis * Transportas vnt.
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        transport_unit_col = get_col_letter_by_header("Transportas vnt. (€)")
        return f"={qty_col}{row_idx}*{transport_unit_col}{row_idx}"
    
    elif header == "Savikaina vnt. be PVM (€)":
        # Vnt. kaina su nuolaida + Akcizas vnt. + Transportas vnt.
        price_discount_col = get_col_letter_by_header("Vnt. kaina su nuolaida (€)")
        excise_unit_col = get_col_letter_by_header("Akcizas vnt. (€)")
        transport_unit_col = get_col_letter_by_header("Transportas vnt. (€)")
        return f"={price_discount_col}{row_idx}+{excise_unit_col}{row_idx}+{transport_unit_col}{row_idx}"
    
    elif header == "Savikaina vnt. su PVM (€)":
        # Savikaina vnt. be PVM * 1.21
        cost_wo_vat_col = get_col_letter_by_header("Savikaina vnt. be PVM (€)")
        return f"={cost_wo_vat_col}{row_idx}*1.21"
    
    elif header == "Savikaina viso be PVM (€)":
        # Kiekis * Savikaina vnt. be PVM
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        cost_wo_vat_col = get_col_letter_by_header("Savikaina vnt. be PVM (€)")
        return f"={qty_col}{row_idx}*{cost_wo_vat_col}{row_idx}"
    
    elif header == "Savikaina viso su PVM (€)":
        # Kiekis * Savikaina vnt. su PVM
        qty_col = get_col_letter_by_header("Kiekis (vnt)")
        cost_w_vat_col = get_col_letter_by_header("Savikaina vnt. su PVM (€)")
        return f"={qty_col}{row_idx}*{cost_w_vat_col}{row_idx}"
    
    return None

def _should_sum_column(header: str) -> bool:
    """Nustato ar stulpelį reikia sumuoti."""
    sum_columns = [
        "Kiekis (vnt)",
        "Suma (€)", 
        "Suma su nuolaida (€)", 
        "Akcizas viso (€)", 
        "Transportas viso (€)", 
        "Savikaina viso be PVM (€)", 
        "Savikaina viso su PVM (€)"
    ]
    return header in sum_columns

def _finalize_excel_formatting(wb, ws, output_path=None):
    """Užbaigia Excel formatavimą ir išsaugo failą."""
    if output_path is None:
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
        filename = f"akcizo_apskaita_{timestamp}.xlsx"
        os.makedirs('output', exist_ok=True)
        output_path = os.path.join('output', filename)
    
    # Stilių nustatymas
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Pritaikome stilius visoms ląstelėms
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            
            # Header eilutės stilius jau nustatytas
            if cell.row == 1:
                continue
            
            # Sumų eilutės stilius
            elif cell.row == ws.max_row and ws.max_row > 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right" if cell.column > 1 else "left", vertical="center")
                if cell.column > 1:  # Ne pirmasis stulpelis
                    cell.number_format = '#,##0.00'
            
            # Duomenų eilučių stilius
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Skaitinių stulpelių formatavimas
                header_text = ws.cell(row=1, column=cell.column).value
                if header_text:
                    if header_text == "Tūris (L)":
                        cell.number_format = '0.000'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif header_text in ["Alk. %", "Nuolaida (%)"]:
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif header_text == "Kiekis (vnt)":
                        cell.number_format = '0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif header_text in ["Akcizas vnt. (€)", "Transportas vnt. (€)", "Transportas viso (€)", "Savikaina vnt. be PVM (€)", "Savikaina vnt. su PVM (€)"]:
                        # Visi € stulpeliai turi būti su 2 skaičiais po kablelio
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif "€" in header_text and header_text != "Akcizo kat. raktas":
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Stulpelių pločio nustatymas
    _adjust_column_widths(ws)
    
    # Header eilutės aukščio nustatymas
    ws.row_dimensions[1].height = 30
    
    wb.save(output_path)
    logging.info("Excel failas su formulėmis sėkmingai sugeneruotas: %s", output_path)
    return output_path

def _adjust_column_widths(ws):
    """Nustato optimalų stulpelių plotį."""
    for col_idx, column_cells in enumerate(ws.columns, 1):
        column_letter = get_column_letter(col_idx)
        header_cell = ws.cell(row=1, column=col_idx)
        header_value = str(header_cell.value) if header_cell.value else ""
        
        # Bazinis plotis pagal header
        if col_idx == 1:  # Prekės pavadinimas
            width = 35
        elif "kategorija" in header_value.lower():
            width = 25
        elif "€" in header_value:
            width = 15
        else:
            width = 12
        
        # Maksimalus plotis
        max_width = 60 if col_idx == 1 else 25
        width = min(width, max_width)
        
        ws.column_dimensions[column_letter].width = width

# Backward compatibility - paliekame senąją funkciją
def generate_excel_file(products_data: list) -> str:
    """Backward compatibility wrapper - naudoja naują funkciją su formulėmis."""
    return generate_excel_file_with_formulas(products_data)